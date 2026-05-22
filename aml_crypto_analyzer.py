#!/usr/bin/env python3
"""
AML Crypto Flow Analyzer

Features:
- Transaction and address deep-dive across multiple chains.
- Flow tracing via graph expansion over linked transactions.
- Mixer detection (known-address + behavioral heuristics).
- Sanctions screening using local watchlists.
- Address ownership hinting from analyst-maintained intelligence files.
- Exports to JSON, XLSX, and GraphML (yEd-compatible).
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import statistics
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

import requests

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None


APP_VERSION = "0.1.0"
TIMEOUT_SECONDS = 20
USER_AGENT = "AML-Crypto-Analyzer/0.1"
DEFAULT_TRANSACTION_LIMIT = 50
COINSTATS_API_ROOT = "https://openapiv1.coinstats.app"
COINSTATS_SYNC_POLL_SECONDS = 2.0
COINSTATS_SYNC_MAX_POLLS = 15
COINSTATS_TRANSACTION_PAGE_SIZE = 100
TRON_ADDRESS_RE = re.compile(r"\bT[1-9A-HJ-NP-Za-km-z]{33}\b")
EVM_ADDRESS_RE = re.compile(r"\b0x[a-fA-F0-9]{40}\b")


SUPPORTED_ASSETS: Dict[str, Dict[str, str]] = {
    "BTC": {"chain_type": "utxo", "blockchair_chain": "bitcoin", "blockcypher_chain": "btc", "blockcypher_net": "main"},
    "ETH": {"chain_type": "account", "blockchair_chain": "ethereum", "blockcypher_chain": "eth", "blockcypher_net": "main"},
    "BNB": {"chain_type": "account", "coinstats_blockchain": "binance_smart"},
    "LTC": {"chain_type": "utxo", "blockchair_chain": "litecoin", "blockcypher_chain": "ltc", "blockcypher_net": "main"},
    "DOGE": {"chain_type": "utxo", "blockchair_chain": "dogecoin", "blockcypher_chain": "doge", "blockcypher_net": "main"},
    "DASH": {"chain_type": "utxo", "blockchair_chain": "dash", "blockcypher_chain": "dash", "blockcypher_net": "main"},
    "BCH": {"chain_type": "utxo", "blockchair_chain": "bitcoin-cash"},
    "XRP": {"chain_type": "account", "blockchair_chain": "ripple"},
    "XLM": {"chain_type": "account", "blockchair_chain": "stellar"},
    "ZEC": {"chain_type": "utxo", "blockchair_chain": "zcash"},
    "XMR": {"chain_type": "account", "blockchair_chain": "monero"},
    "TRX": {"chain_type": "account", "coinstats_connection_id": "tron"},
}

ASSET_COINGECKO_IDS: Dict[str, str] = {
    "BTC": "bitcoin",
    "ETH": "ethereum",
    "BNB": "binancecoin",
    "LTC": "litecoin",
    "DOGE": "dogecoin",
    "DASH": "dash",
    "BCH": "bitcoin-cash",
    "XRP": "ripple",
    "XLM": "stellar",
    "ZEC": "zcash",
    "XMR": "monero",
    "TRX": "tron",
}

ASSET_COINSTATS_COIN_IDS: Dict[str, str] = {
    "TRX": "tron",
}


def normalize_lookup_address(address: str) -> str:
    value = str(address or "").strip()
    # EVM addresses are case-insensitive for matching purposes.
    if value.lower().startswith("0x") and len(value) == 42:
        return value.lower()
    return value


@dataclass
class TxIO:
    address: str
    amount: float


@dataclass
class TransactionRecord:
    asset: str
    txid: str
    timestamp: Optional[str]
    inputs: List[TxIO]
    outputs: List[TxIO]
    fee: Optional[float] = None
    raw: Dict[str, Any] = field(default_factory=dict)


@dataclass
class GraphEdge:
    src: str
    dst: str
    label: str
    amount: float
    txid: str
    asset: str


@dataclass
class Finding:
    level: str
    finding_type: str
    subject: str
    detail: str


class ProviderError(RuntimeError):
    pass


class HttpClient:
    def __init__(self) -> None:
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})

    def get_json(
        self,
        url: str,
        params: Optional[Dict[str, Any]] = None,
        headers: Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        response = self.session.get(url, params=params or {}, headers=headers or {}, timeout=TIMEOUT_SECONDS)
        if response.status_code >= 400:
            raise ProviderError(f"HTTP {response.status_code} from {url}: {response.text[:300]}")
        return response.json()

    def patch_json(
        self,
        url: str,
        params: Optional[Dict[str, Any]] = None,
        payload: Optional[Dict[str, Any]] = None,
        headers: Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        response = self.session.patch(
            url,
            params=params or {},
            json=payload or {},
            headers=headers or {},
            timeout=TIMEOUT_SECONDS,
        )
        if response.status_code >= 400:
            raise ProviderError(f"HTTP {response.status_code} from {url}: {response.text[:300]}")
        return response.json()


class BlockCypherProvider:
    def __init__(self, api_key: Optional[str], http: HttpClient) -> None:
        self.api_key = api_key
        self.http = http

    def supports(self, asset: str) -> bool:
        info = SUPPORTED_ASSETS.get(asset, {})
        return "blockcypher_chain" in info

    def _base(self, asset: str) -> str:
        info = SUPPORTED_ASSETS[asset]
        return f"https://api.blockcypher.com/v1/{info['blockcypher_chain']}/{info['blockcypher_net']}"

    def fetch_transaction(self, asset: str, txid: str) -> TransactionRecord:
        if not self.supports(asset):
            raise ProviderError(f"BlockCypher does not support {asset}")
        params: Dict[str, Any] = {"limit": 2000}
        if self.api_key:
            params["token"] = self.api_key
        payload = self.http.get_json(f"{self._base(asset)}/txs/{txid}", params=params)

        unit_divisor = self._unit_divisor(asset)

        inputs: List[TxIO] = []
        for txin in payload.get("inputs", []):
            addresses = txin.get("addresses", [])
            value = float(txin.get("output_value", txin.get("value", 0))) / unit_divisor
            for addr in addresses:
                inputs.append(TxIO(address=addr, amount=value))

        outputs: List[TxIO] = []
        for txout in payload.get("outputs", []):
            addresses = txout.get("addresses", [])
            value = float(txout.get("value", 0)) / unit_divisor
            for addr in addresses:
                outputs.append(TxIO(address=addr, amount=value))

        # Account-style chains (e.g., ETH) may not expose UTXO-style inputs/outputs.
        if not inputs and not outputs:
            from_addr = payload.get("from")
            to_addr = payload.get("to")
            value = float(payload.get("value", 0)) / unit_divisor
            if from_addr:
                inputs.append(TxIO(address=str(from_addr), amount=value))
            if to_addr:
                outputs.append(TxIO(address=str(to_addr), amount=value))

        fee = float(payload.get("fees", payload.get("gas_used", 0))) / unit_divisor if payload.get("fees") is not None or payload.get("gas_used") is not None else None
        return TransactionRecord(
            asset=asset,
            txid=payload.get("hash", txid),
            timestamp=payload.get("confirmed"),
            inputs=inputs,
            outputs=outputs,
            fee=fee,
            raw=payload,
        )

    def fetch_address_transactions(self, asset: str, address: str, limit: Optional[int] = 10) -> List[str]:
        if not self.supports(asset):
            return []
        provider_limit = 50 if limit is None else max(1, min(limit, 50))
        params: Dict[str, Any] = {"limit": provider_limit, "txlimit": provider_limit}
        if self.api_key:
            params["token"] = self.api_key
        payload = self.http.get_json(f"{self._base(asset)}/addrs/{address}", params=params)
        txrefs = payload.get("txrefs", []) + payload.get("unconfirmed_txrefs", [])
        txids = [r.get("tx_hash") for r in txrefs if r.get("tx_hash")]
        return txids if limit is None else txids[:limit]

    @staticmethod
    def _unit_divisor(asset: str) -> float:
        # UTXO chains use satoshi-like base units. ETH uses wei.
        if asset == "ETH":
            return 1e18
        return 1e8


class BlockchairProvider:
    def __init__(self, http: HttpClient, api_key: Optional[str] = None) -> None:
        self.http = http
        self.api_key = api_key

    def supports(self, asset: str) -> bool:
        info = SUPPORTED_ASSETS.get(asset, {})
        return "blockchair_chain" in info

    def _chain(self, asset: str) -> str:
        return SUPPORTED_ASSETS[asset]["blockchair_chain"]

    def fetch_transaction(self, asset: str, txid: str) -> TransactionRecord:
        if not self.supports(asset):
            raise ProviderError(f"Blockchair does not support {asset}")
        url = f"https://api.blockchair.com/{self._chain(asset)}/dashboards/transaction/{txid}"
        params: Dict[str, Any] = {}
        if self.api_key:
            params["key"] = self.api_key
        payload = self.http.get_json(url, params=params)
        data = payload.get("data", {})
        tx_data = data.get(txid, {}) if isinstance(data, dict) else {}

        raw_inputs = tx_data.get("inputs", []) if isinstance(tx_data, dict) else []
        raw_outputs = tx_data.get("outputs", []) if isinstance(tx_data, dict) else []
        tx_meta = tx_data.get("transaction", {}) if isinstance(tx_data, dict) else {}

        inputs: List[TxIO] = []
        for txin in raw_inputs:
            addr = str(txin.get("recipient") or txin.get("sender") or txin.get("address") or "unknown")
            val = txin.get("value", txin.get("amount", 0))
            amount = self._to_float(val)
            inputs.append(TxIO(address=addr, amount=amount))

        outputs: List[TxIO] = []
        for txout in raw_outputs:
            addr = str(txout.get("recipient") or txout.get("sender") or txout.get("address") or "unknown")
            val = txout.get("value", txout.get("amount", 0))
            amount = self._to_float(val)
            outputs.append(TxIO(address=addr, amount=amount))

        timestamp = tx_meta.get("time") or tx_meta.get("block_time")
        fee = self._to_float(tx_meta.get("fee", 0)) if tx_meta else None

        if not inputs and not outputs:
            raise ProviderError(
                f"Could not normalize transaction {txid} for {asset}; explorer schema may have changed."
            )

        return TransactionRecord(
            asset=asset,
            txid=txid,
            timestamp=timestamp,
            inputs=inputs,
            outputs=outputs,
            fee=fee,
            raw=payload,
        )

    def fetch_address_transactions(self, asset: str, address: str, limit: Optional[int] = 10) -> List[str]:
        if not self.supports(asset):
            return []
        url = f"https://api.blockchair.com/{self._chain(asset)}/dashboards/address/{address}"
        params: Dict[str, Any] = {}
        if self.api_key:
            params["key"] = self.api_key
        payload = self.http.get_json(url, params=params)
        data = payload.get("data", {})
        addr_data = self._select_address_data(data, address)
        if not isinstance(addr_data, dict):
            return []

        txids: List[str] = []
        # Different chains expose different fields in Blockchair dashboards.
        for key in ("transactions", "calls", "token_transfers"):
            txids.extend(self._normalize_txids(addr_data.get(key, [])))

        deduped: List[str] = []
        seen: Set[str] = set()
        for txid in txids:
            if txid not in seen:
                deduped.append(txid)
                seen.add(txid)
            if limit is not None and len(deduped) >= limit:
                break
        return deduped

    @staticmethod
    def _to_float(value: Any) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            if value > 1_000_000_000:
                return float(value) / 1e8
            return float(value)
        if isinstance(value, str):
            try:
                return float(value)
            except ValueError:
                return 0.0
        return 0.0

    @staticmethod
    def _select_address_data(data: Any, address: str) -> Dict[str, Any]:
        if not isinstance(data, dict):
            return {}
        if address in data and isinstance(data[address], dict):
            return data[address]
        address_lc = address.lower()
        for key, value in data.items():
            if str(key).lower() == address_lc and isinstance(value, dict):
                return value
        # Fallback: some responses include a single address-like object keyed unexpectedly.
        for value in data.values():
            if isinstance(value, dict):
                return value
        return {}

    @staticmethod
    def _normalize_txids(raw_items: Any) -> List[str]:
        if not isinstance(raw_items, list):
            return []
        out: List[str] = []
        for item in raw_items:
            if isinstance(item, (str, int)):
                out.append(str(item))
                continue
            if isinstance(item, dict):
                for key in ("transaction_hash", "tx_hash", "hash", "id"):
                    value = item.get(key)
                    if value:
                        out.append(str(value))
                        break
        return out


class CoinStatsProvider:
    def __init__(self, http: HttpClient, api_key: Optional[str] = None) -> None:
        self.http = http
        self.api_key = api_key
        self._tx_cache: Dict[Tuple[str, str], TransactionRecord] = {}
        self._tx_wallet_map: Dict[Tuple[str, str], str] = {}
        self._synced_wallets: Set[Tuple[str, str]] = set()

    def supports(self, asset: str) -> bool:
        info = SUPPORTED_ASSETS.get(asset, {})
        return "coinstats_connection_id" in info or "coinstats_blockchain" in info

    def _connection_id(self, asset: str) -> str:
        return str(SUPPORTED_ASSETS[asset]["coinstats_connection_id"])

    def _query_params(self, asset: str, address: str) -> Dict[str, Any]:
        info = SUPPORTED_ASSETS[asset]
        params: Dict[str, Any] = {"address": address}
        if "coinstats_connection_id" in info:
            params["connectionId"] = str(info["coinstats_connection_id"])
        elif "coinstats_blockchain" in info:
            params["blockchain"] = str(info["coinstats_blockchain"])
        else:
            raise ProviderError(f"CoinStats mapping not configured for {asset}")
        return params

    @staticmethod
    def _coin_params(asset: str) -> Dict[str, Any]:
        coin_id = ASSET_COINSTATS_COIN_IDS.get(asset)
        return {"coinId": coin_id} if coin_id else {}

    def _transaction_query_params(self, asset: str, address: str) -> Dict[str, Any]:
        params = self._query_params(asset, address)
        params.update(self._coin_params(asset))
        return params

    def _headers(self) -> Dict[str, str]:
        if not self.api_key:
            raise ProviderError("CoinStats API key is required for this asset")
        return {"X-API-KEY": self.api_key}

    def fetch_transaction(
        self,
        asset: str,
        txid: str,
        wallet_address: Optional[str] = None,
    ) -> TransactionRecord:
        if not self.supports(asset):
            raise ProviderError(f"CoinStats does not support {asset}")
        cache_key = (asset, txid)
        if cache_key in self._tx_cache:
            return self._tx_cache[cache_key]

        params: Dict[str, Any] = {
            "txId": txid,
            "page": 1,
            "limit": 1,
        }
        params.update(self._coin_params(asset))
        query_address = wallet_address or self._tx_wallet_map.get(cache_key)
        if query_address:
            params.update(self._transaction_query_params(asset, query_address))
            self._ensure_wallet_synced(asset=asset, address=query_address)

        payload = self.http.get_json(
            f"{COINSTATS_API_ROOT}/wallet/transactions",
            params=params,
            headers=self._headers(),
        )
        result = payload.get("result", [])
        if not isinstance(result, list):
            raise ProviderError(f"CoinStats returned unexpected transaction payload for {txid}")

        for entry in result:
            record = self._normalize_transaction(asset=asset, entry=entry, wallet_address=query_address)
            if self._txid_matches(record.txid, txid):
                return record

        raise ProviderError(
            f"CoinStats could not resolve transaction {txid} for {asset}. "
            f"Try tracing from a supported wallet address or ensure the wallet is synced."
        )

    def fetch_address_transactions(self, asset: str, address: str, limit: Optional[int] = 10) -> List[str]:
        if not self.supports(asset):
            return []
        self._ensure_wallet_synced(asset=asset, address=address)

        txids: List[str] = []
        seen: Set[str] = set()
        per_page = (
            COINSTATS_TRANSACTION_PAGE_SIZE
            if limit is None
            else max(1, min(COINSTATS_TRANSACTION_PAGE_SIZE, limit))
        )
        page = 1
        while limit is None or len(txids) < limit:
            payload = self.http.get_json(
                f"{COINSTATS_API_ROOT}/wallet/transactions",
                params={**self._transaction_query_params(asset, address), "page": page, "limit": per_page},
                headers=self._headers(),
            )
            result = payload.get("result", [])
            if not isinstance(result, list) or not result:
                break

            new_count = 0
            for entry in result:
                record = self._normalize_transaction(asset=asset, entry=entry, wallet_address=address)
                if record.txid and record.txid not in seen:
                    txids.append(record.txid)
                    seen.add(record.txid)
                    new_count += 1
                if limit is not None and len(txids) >= limit:
                    break

            if new_count == 0 or len(result) < per_page:
                break
            page += 1

        return txids if limit is None else txids[:limit]

    def _ensure_wallet_synced(self, asset: str, address: str) -> None:
        wallet_key = (asset, address)
        if wallet_key in self._synced_wallets:
            return

        params = self._query_params(asset, address)
        self.http.patch_json(
            f"{COINSTATS_API_ROOT}/wallet/transactions",
            params=params,
            headers=self._headers(),
        )

        last_status = "unknown"
        for _ in range(COINSTATS_SYNC_MAX_POLLS):
            payload = self.http.get_json(
                f"{COINSTATS_API_ROOT}/wallet/status",
                params=params,
                headers=self._headers(),
            )
            last_status = str(payload.get("status", "")).strip().lower()
            if last_status == "synced":
                self._synced_wallets.add(wallet_key)
                return
            time.sleep(COINSTATS_SYNC_POLL_SECONDS)

        raise ProviderError(f"CoinStats wallet sync did not complete for {address} on {asset} (last status: {last_status})")

    def _normalize_transaction(self, asset: str, entry: Dict[str, Any], wallet_address: Optional[str]) -> TransactionRecord:
        txid = self._extract_hash(entry)
        timestamp = entry.get("date")
        wallet_norm = normalize_lookup_address(wallet_address or "")
        from_addresses = self._extract_addresses(entry.get("fromAddress"))
        to_addresses = self._extract_addresses(entry.get("toAddress"))
        discovered = self._extract_addresses(entry)
        counterparties = [addr for addr in discovered if normalize_lookup_address(addr) != wallet_norm]
        total_amount = self._extract_amount(entry)
        tx_type = str(entry.get("type", "")).strip().lower()

        inputs: List[TxIO] = []
        outputs: List[TxIO] = []

        def split_targets(addresses: List[str], amount: float) -> List[TxIO]:
            if not addresses or amount <= 0:
                return []
            share = amount / len(addresses)
            return [TxIO(address=a, amount=share) for a in addresses]

        is_outbound = total_amount < 0 or tx_type in {"sent", "withdraw", "approve", "executed", "fee"}
        amount_abs = abs(total_amount)

        if amount_abs > 0 and (from_addresses or to_addresses):
            if from_addresses:
                inputs.extend(split_targets(from_addresses, amount_abs))
            elif wallet_address and is_outbound:
                inputs.append(TxIO(address=wallet_address, amount=amount_abs))

            if to_addresses:
                outputs.extend(split_targets(to_addresses, amount_abs))
            elif wallet_address and not is_outbound:
                outputs.append(TxIO(address=wallet_address, amount=amount_abs))
        elif wallet_address:
            if is_outbound:
                if amount_abs > 0:
                    inputs.append(TxIO(address=wallet_address, amount=amount_abs))
                outputs.extend(split_targets(counterparties, amount_abs))
            else:
                inputs.extend(split_targets(counterparties, amount_abs))
                if amount_abs > 0:
                    outputs.append(TxIO(address=wallet_address, amount=amount_abs))
        else:
            if discovered and amount_abs > 0:
                if len(discovered) == 1:
                    target = discovered[0]
                    if is_outbound:
                        inputs.append(TxIO(address=target, amount=amount_abs))
                    else:
                        outputs.append(TxIO(address=target, amount=amount_abs))
                else:
                    inputs.append(TxIO(address=discovered[0], amount=amount_abs))
                    outputs.extend(split_targets(discovered[1:], amount_abs))

        fee = self._extract_fee(entry)
        record = TransactionRecord(
            asset=asset,
            txid=txid,
            timestamp=str(timestamp) if timestamp else None,
            inputs=inputs,
            outputs=outputs,
            fee=fee,
            raw=entry,
        )
        self._tx_cache[(asset, txid)] = record
        if wallet_address:
            self._tx_wallet_map[(asset, txid)] = wallet_address
        return record

    @staticmethod
    def _extract_hash(entry: Dict[str, Any]) -> str:
        hash_data = entry.get("hash", {})
        if isinstance(hash_data, dict) and hash_data.get("id"):
            return str(hash_data.get("id"))
        for key in ("txid", "transactionHash", "hash", "id"):
            value = entry.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
        raise ProviderError("CoinStats transaction payload missing hash/id")

    @staticmethod
    def _txid_matches(candidate: str, expected: str) -> bool:
        return str(candidate).strip().lower() == str(expected).strip().lower()

    @staticmethod
    def _extract_amount(entry: Dict[str, Any]) -> float:
        coin_data = entry.get("coinData", {})
        if isinstance(coin_data, dict) and coin_data.get("count") is not None:
            try:
                return float(coin_data.get("count"))
            except Exception:
                pass

        total = 0.0
        found = False
        for tx in entry.get("transactions", []) if isinstance(entry.get("transactions"), list) else []:
            items = tx.get("items", []) if isinstance(tx, dict) else []
            for item in items:
                if not isinstance(item, dict) or item.get("count") is None:
                    continue
                try:
                    total += float(item.get("count"))
                    found = True
                except Exception:
                    continue
        return total if found else 0.0

    @staticmethod
    def _extract_fee(entry: Dict[str, Any]) -> Optional[float]:
        fee = entry.get("fee", {})
        if not isinstance(fee, dict) or fee.get("count") is None:
            return None
        try:
            return abs(float(fee.get("count")))
        except Exception:
            return None

    @staticmethod
    def _extract_addresses(payload: Any) -> List[str]:
        found: List[str] = []
        seen: Set[str] = set()

        def add_match(value: str) -> None:
            normalized = normalize_lookup_address(value)
            if normalized not in seen:
                seen.add(normalized)
                found.append(normalized)

        def walk(node: Any) -> None:
            if isinstance(node, str):
                for match in TRON_ADDRESS_RE.findall(node):
                    add_match(match)
                for match in EVM_ADDRESS_RE.findall(node):
                    add_match(match)
                return
            if isinstance(node, list):
                for item in node:
                    walk(item)
                return
            if isinstance(node, dict):
                for value in node.values():
                    walk(value)

        walk(payload)
        return found


class ProviderRouter:
    def __init__(
        self,
        blockcypher_key: Optional[str],
        blockchair_key: Optional[str] = None,
        coinstats_key: Optional[str] = None,
        allow_blockchair_fallback: bool = False,
    ) -> None:
        http = HttpClient()
        self.blockcypher = BlockCypherProvider(api_key=blockcypher_key, http=http)
        self.blockchair = BlockchairProvider(http=http, api_key=blockchair_key)
        self.coinstats = CoinStatsProvider(http=http, api_key=coinstats_key)
        self.allow_blockchair_fallback = allow_blockchair_fallback

    def fetch_transaction(self, asset: str, txid: str) -> TransactionRecord:
        errors: List[str] = []
        blockcypher_supported = self.blockcypher.supports(asset)
        blockchair_supported = self.blockchair.supports(asset)
        coinstats_supported = self.coinstats.supports(asset)

        if blockcypher_supported:
            try:
                return self.blockcypher.fetch_transaction(asset, txid)
            except Exception as exc:
                errors.append(f"BlockCypher: {exc}")
        if blockchair_supported:
            # Blockchair is strict fallback only: used when BlockCypher failed
            # or for assets BlockCypher does not support.
            if (not blockcypher_supported) or (self.allow_blockchair_fallback and errors):
                try:
                    return self.blockchair.fetch_transaction(asset, txid)
                except Exception as exc:
                    errors.append(f"Blockchair: {exc}")
        if coinstats_supported:
            try:
                return self.coinstats.fetch_transaction(asset, txid)
            except Exception as exc:
                errors.append(f"CoinStats: {exc}")
        raise ProviderError("; ".join(errors) if errors else f"No provider available for {asset}")

    def fetch_address_transactions(
        self,
        asset: str,
        address: str,
        limit: Optional[int] = DEFAULT_TRANSACTION_LIMIT,
    ) -> List[str]:
        errors: List[str] = []
        blockcypher_supported = self.blockcypher.supports(asset)
        blockchair_supported = self.blockchair.supports(asset)
        coinstats_supported = self.coinstats.supports(asset)

        if blockcypher_supported:
            try:
                txids = self.blockcypher.fetch_address_transactions(asset, address, limit=limit)
                if txids:
                    return txids if limit is None else txids[:limit]
                errors.append("BlockCypher: returned no transactions")
            except Exception as exc:
                errors.append(f"BlockCypher: {exc}")

        if blockchair_supported and ((not blockcypher_supported) or (self.allow_blockchair_fallback and errors)):
            try:
                txids = self.blockchair.fetch_address_transactions(asset, address, limit=limit)
                if txids:
                    return txids if limit is None else txids[:limit]
                errors.append("Blockchair: returned no transactions")
            except Exception as exc:
                errors.append(f"Blockchair: {exc}")
        if coinstats_supported:
            try:
                txids = self.coinstats.fetch_address_transactions(asset, address, limit=limit)
                if txids:
                    return txids if limit is None else txids[:limit]
                errors.append("CoinStats: returned no transactions")
            except Exception as exc:
                errors.append(f"CoinStats: {exc}")

        if errors:
            raise ProviderError(f"Address lookup failed for {address} on {asset}: " + "; ".join(errors))
        raise ProviderError(f"No provider available for {asset} address lookup")


class IntelligenceStore:
    def __init__(self, intel_dir: Path) -> None:
        self.intel_dir = intel_dir
        self.sanctioned: Dict[str, str] = self._load_map_csv("sanctioned_addresses.csv", ["address", "source"])
        self.mixers: Dict[str, str] = self._load_map_csv("known_mixers.csv", ["address", "mixer_name"])
        self.owners: Dict[str, str] = self._load_map_csv("known_entities.csv", ["address", "entity"])

    def _load_map_csv(self, filename: str, required: List[str]) -> Dict[str, str]:
        path = self.intel_dir / filename
        if not path.exists():
            return {}
        data: Dict[str, str] = {}
        # Use utf-8-sig to tolerate CSVs saved with BOM by some Windows editors.
        with path.open("r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for field in required:
                if field not in reader.fieldnames:
                    raise ValueError(f"{path} missing required field: {field}")
            for row in reader:
                addr = str(row[required[0]]).strip()
                value = str(row[required[1]]).strip()
                if addr:
                    data[addr] = value
                    normalized = normalize_lookup_address(addr)
                    if normalized and normalized not in data:
                        data[normalized] = value
        return data


class Analyzer:
    def __init__(self, provider: ProviderRouter, intel: IntelligenceStore) -> None:
        self.provider = provider
        self.intel = intel

    def trace_from_transaction(
        self,
        asset: str,
        txid: str,
        max_transactions: Optional[int] = DEFAULT_TRANSACTION_LIMIT,
    ) -> Dict[str, Any]:
        max_transactions = self._normalize_transaction_limit(max_transactions)
        visited_txs: Set[str] = set()
        queued_txs: Set[str] = {txid}
        transactions: List[TransactionRecord] = []
        edges: List[GraphEdge] = []
        findings: List[Finding] = []

        queue: List[Tuple[str, int]] = [(txid, 0)]

        while queue and (max_transactions is None or len(transactions) < max_transactions):
            current_txid, level = queue.pop(0)
            queued_txs.discard(current_txid)
            if current_txid in visited_txs:
                continue

            try:
                tx = self.provider.fetch_transaction(asset, current_txid)
            except Exception as exc:
                findings.append(Finding("medium", "trace_gap", current_txid, str(exc)))
                visited_txs.add(current_txid)
                continue

            visited_txs.add(current_txid)
            transactions.append(tx)
            next_addresses = self._append_transaction(tx, hop=level, edges=edges, findings=findings)

            remaining = None if max_transactions is None else max_transactions - len(transactions)
            if remaining is not None and remaining <= 0:
                continue

            for address in next_addresses:
                lookup_limit = remaining
                if lookup_limit is not None and lookup_limit <= 0:
                    break
                child_txids = self.provider.fetch_address_transactions(asset, address, limit=lookup_limit)
                for child in child_txids:
                    if child != tx.txid and child not in visited_txs and child not in queued_txs:
                        queue.append((child, level + 1))
                        queued_txs.add(child)
                        if remaining is not None:
                            remaining -= 1
                            if remaining <= 0:
                                break

        return self._build_result(asset, transactions, edges, findings, transaction_limit=max_transactions)

    def trace_from_address(
        self,
        asset: str,
        address: str,
        max_transactions: Optional[int] = DEFAULT_TRANSACTION_LIMIT,
    ) -> Dict[str, Any]:
        max_transactions = self._normalize_transaction_limit(max_transactions)
        txids = self.provider.fetch_address_transactions(asset, address, limit=max_transactions)
        if not txids:
            raise ProviderError(f"No transactions found for {address} on {asset}")
        visited_txs: Set[str] = set()
        transactions: List[TransactionRecord] = []
        edges: List[GraphEdge] = []
        findings: List[Finding] = []

        for txid in txids:
            if txid in visited_txs:
                continue
            if max_transactions is not None and len(transactions) >= max_transactions:
                break

            try:
                tx = self.provider.fetch_transaction(asset, txid)
            except Exception as exc:
                findings.append(Finding("medium", "trace_gap", txid, str(exc)))
                visited_txs.add(txid)
                continue

            visited_txs.add(txid)
            transactions.append(tx)
            self._append_transaction(tx, hop=0, edges=edges, findings=findings)

        return self._build_result(asset, transactions, edges, findings, transaction_limit=max_transactions)

    @staticmethod
    def _normalize_transaction_limit(max_transactions: Optional[int]) -> Optional[int]:
        if max_transactions is not None and max_transactions < 1:
            raise ValueError("max_transactions must be >= 1 or None for all transactions")
        return max_transactions

    def _append_transaction(
        self,
        tx: TransactionRecord,
        hop: int,
        edges: List[GraphEdge],
        findings: List[Finding],
    ) -> List[str]:
        findings.extend(self._analyze_transaction(tx, hop=hop))

        for txin in tx.inputs:
            edges.append(
                GraphEdge(
                    src=f"addr:{txin.address}",
                    dst=f"tx:{tx.txid}",
                    label="input",
                    amount=txin.amount,
                    txid=tx.txid,
                    asset=tx.asset,
                )
            )

        next_addresses: List[str] = []
        for txout in tx.outputs:
            edges.append(
                GraphEdge(
                    src=f"tx:{tx.txid}",
                    dst=f"addr:{txout.address}",
                    label="output",
                    amount=txout.amount,
                    txid=tx.txid,
                    asset=tx.asset,
                )
            )
            next_addresses.append(txout.address)

        return next_addresses

    def _analyze_transaction(self, tx: TransactionRecord, hop: int) -> List[Finding]:
        findings: List[Finding] = []

        for i in tx.inputs + tx.outputs:
            addr = str(i.address).strip()
            addr_norm = normalize_lookup_address(addr)
            sanctioned_source = self.intel.sanctioned.get(addr) or self.intel.sanctioned.get(addr_norm)
            mixer_name = self.intel.mixers.get(addr) or self.intel.mixers.get(addr_norm)
            owner_name = self.intel.owners.get(addr) or self.intel.owners.get(addr_norm)

            if sanctioned_source:
                findings.append(
                    Finding(
                        level="high",
                        finding_type="sanctioned_address",
                        subject=addr,
                        detail=f"Listed in sanctions source: {sanctioned_source}",
                    )
                )
            if mixer_name:
                findings.append(
                    Finding(
                        level="high",
                        finding_type="known_mixer",
                        subject=addr,
                        detail=f"Known mixer service: {mixer_name}",
                    )
                )
            if owner_name:
                findings.append(
                    Finding(
                        level="info",
                        finding_type="owner_hint",
                        subject=addr,
                        detail=f"Analyst-tagged entity: {owner_name}",
                    )
                )

        if len(tx.outputs) >= 8:
            amounts = [o.amount for o in tx.outputs if o.amount > 0]
            if len(amounts) >= 8:
                stdev = statistics.pstdev(amounts) if len(amounts) > 1 else 0.0
                mean = statistics.mean(amounts) if amounts else 0.0
                rel_var = (stdev / mean) if mean > 0 else math.inf
                if rel_var < 0.15 and hop <= 2:
                    findings.append(
                        Finding(
                            level="medium",
                            finding_type="mixer_pattern",
                            subject=tx.txid,
                            detail="High fan-out with near-uniform output values (possible peel chain/mixer behavior)",
                        )
                    )

        return findings

    def _build_result(
        self,
        asset: str,
        transactions: List[TransactionRecord],
        edges: List[GraphEdge],
        findings: List[Finding],
        transaction_limit: Optional[int] = DEFAULT_TRANSACTION_LIMIT,
    ) -> Dict[str, Any]:
        nodes: Dict[str, Dict[str, Any]] = {}
        for tx in transactions:
            nodes[f"tx:{tx.txid}"] = {"id": f"tx:{tx.txid}", "type": "transaction", "asset": tx.asset}
            for item in tx.inputs + tx.outputs:
                address_node = f"addr:{item.address}"
                if address_node not in nodes:
                    nodes[address_node] = {"id": address_node, "type": "address", "asset": tx.asset}

        return {
            "metadata": {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "asset": asset,
                "version": APP_VERSION,
                "transaction_limit": "all" if transaction_limit is None else transaction_limit,
                "transaction_count": len(transactions),
                "node_count": len(nodes),
                "edge_count": len(edges),
                "finding_count": len(findings),
            },
            "transactions": [self._tx_to_dict(t) for t in transactions],
            "graph": {
                "nodes": list(nodes.values()),
                "edges": [e.__dict__ for e in edges],
            },
            "findings": [f.__dict__ for f in findings],
        }

    @staticmethod
    def _tx_to_dict(tx: TransactionRecord) -> Dict[str, Any]:
        return {
            "asset": tx.asset,
            "txid": tx.txid,
            "timestamp": tx.timestamp,
            "fee": tx.fee,
            "inputs": [io.__dict__ for io in tx.inputs],
            "outputs": [io.__dict__ for io in tx.outputs],
        }


class HistoricalFxConverter:
    def __init__(self, http: Optional[HttpClient] = None) -> None:
        self.http = http or HttpClient()
        self._cache: Dict[Tuple[str, str, str], Tuple[float, str]] = {}

    def get_rate(self, asset: str, timestamp: Optional[str], fiat_currency: str) -> Tuple[float, str]:
        symbol = asset.upper().strip()
        currency = fiat_currency.upper().strip()
        date_str = self._extract_date(timestamp)
        cache_key = (symbol, date_str, currency)
        if cache_key in self._cache:
            return self._cache[cache_key]

        coin_id = ASSET_COINGECKO_IDS.get(symbol)
        if not coin_id:
            raise ProviderError(f"No fiat conversion mapping configured for asset {symbol}")

        day, month, year = date_str.split("-")
        coingecko_date = f"{day}-{month}-{year}"
        url = f"https://api.coingecko.com/api/v3/coins/{coin_id}/history"
        payload = self.http.get_json(url, params={"date": coingecko_date, "localization": "false"})
        market = payload.get("market_data", {}).get("current_price", {})
        rate = market.get(currency.lower())
        if rate is None:
            raise ProviderError(f"No {currency} historical rate for {symbol} on {date_str}")

        value = float(rate)
        source = f"CoinGecko {coin_id} history {date_str}"
        self._cache[cache_key] = (value, source)
        return value, source

    @staticmethod
    def _extract_date(timestamp: Optional[str]) -> str:
        if not timestamp:
            return datetime.now(timezone.utc).strftime("%d-%m-%Y")
        ts = str(timestamp).strip()
        if len(ts) >= 10 and ts[4] == "-" and ts[7] == "-":
            dt = ts[:10]
            yyyy, mm, dd = dt.split("-")
            return f"{dd}-{mm}-{yyyy}"
        try:
            normalized = ts.replace("Z", "+00:00")
            parsed = datetime.fromisoformat(normalized)
            return parsed.strftime("%d-%m-%Y")
        except Exception:
            return datetime.now(timezone.utc).strftime("%d-%m-%Y")


def enrich_result_with_fiat(
    result: Dict[str, Any],
    fiat_currency: Optional[str],
    converter: Optional[HistoricalFxConverter] = None,
) -> Dict[str, Any]:
    if not fiat_currency:
        return result
    currency = fiat_currency.upper().strip()
    if not currency or currency == "NONE":
        return result

    conv = converter or HistoricalFxConverter()
    failures: List[str] = []

    for tx in result.get("transactions", []):
        asset = str(tx.get("asset", "")).upper()
        timestamp = tx.get("timestamp")
        try:
            rate, source = conv.get_rate(asset=asset, timestamp=timestamp, fiat_currency=currency)
        except Exception as exc:
            failures.append(f"{tx.get('txid', 'unknown')}: {exc}")
            continue

        total_in = float(sum(float(i.get("amount", 0) or 0) for i in tx.get("inputs", [])))
        total_out = float(sum(float(o.get("amount", 0) or 0) for o in tx.get("outputs", [])))
        fee = tx.get("fee")
        fee_val = float(fee) if fee is not None else None

        tx["fiat_currency"] = currency
        tx["fiat_rate"] = rate
        tx["fiat_rate_source"] = source
        tx["fiat_input_total"] = round(total_in * rate, 2)
        tx["fiat_output_total"] = round(total_out * rate, 2)
        tx["fiat_fee"] = round(fee_val * rate, 2) if fee_val is not None else None

    meta = result.setdefault("metadata", {})
    meta["fiat_currency"] = currency
    if failures:
        meta["fiat_conversion_failures"] = failures[:50]

    return result


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def export_json(result: Dict[str, Any], output_path: Path) -> None:
    ensure_parent(output_path)
    output_path.write_text(json.dumps(result, indent=2), encoding="utf-8")


def export_xlsx(result: Dict[str, Any], output_path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Install dependencies from requirements.txt")

    ensure_parent(output_path)
    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = "metadata"
    ws_meta.append(["key", "value"])
    for k, v in result.get("metadata", {}).items():
        ws_meta.append([k, str(v)])

    ws_tx = wb.create_sheet("transactions")
    ws_tx.append(
        [
            "asset",
            "txid",
            "timestamp",
            "fee",
            "input_count",
            "output_count",
            "fiat_currency",
            "fiat_rate",
            "fiat_input_total",
            "fiat_output_total",
            "fiat_fee",
            "fiat_rate_source",
        ]
    )
    for tx in result.get("transactions", []):
        ws_tx.append([
            tx.get("asset"),
            tx.get("txid"),
            tx.get("timestamp"),
            tx.get("fee"),
            len(tx.get("inputs", [])),
            len(tx.get("outputs", [])),
            tx.get("fiat_currency"),
            tx.get("fiat_rate"),
            tx.get("fiat_input_total"),
            tx.get("fiat_output_total"),
            tx.get("fiat_fee"),
            tx.get("fiat_rate_source"),
        ])

    ws_edges = wb.create_sheet("edges")
    ws_edges.append(["src", "dst", "label", "amount", "txid", "asset"])
    for edge in result.get("graph", {}).get("edges", []):
        ws_edges.append([
            edge.get("src"),
            edge.get("dst"),
            edge.get("label"),
            edge.get("amount"),
            edge.get("txid"),
            edge.get("asset"),
        ])

    ws_find = wb.create_sheet("findings")
    ws_find.append(["level", "finding_type", "subject", "detail"])
    for finding in result.get("findings", []):
        ws_find.append([
            finding.get("level"),
            finding.get("finding_type"),
            finding.get("subject"),
            finding.get("detail"),
        ])

    wb.save(output_path)


def xml_escape(value: Any) -> str:
    s = str(value)
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def export_graphml(result: Dict[str, Any], output_path: Path) -> None:
    ensure_parent(output_path)

    lines: List[str] = []
    lines.append("<?xml version=\"1.0\" encoding=\"UTF-8\"?>")
    lines.append(
        "<graphml xmlns=\"http://graphml.graphdrawing.org/xmlns\" "
        "xmlns:y=\"http://www.yworks.com/xml/graphml\">"
    )
    lines.append('  <key id="d0" for="node" attr.name="type" attr.type="string"/>')
    lines.append('  <key id="d1" for="node" attr.name="asset" attr.type="string"/>')
    lines.append('  <key id="d6" for="node" attr.name="display_label" attr.type="string"/>')
    lines.append('  <key id="d2" for="edge" attr.name="label" attr.type="string"/>')
    lines.append('  <key id="d3" for="edge" attr.name="amount" attr.type="double"/>')
    lines.append('  <key id="d4" for="edge" attr.name="txid" attr.type="string"/>')
    lines.append('  <key id="d5" for="edge" attr.name="asset" attr.type="string"/>')
    lines.append('  <key id="d7" for="edge" attr.name="amount_label" attr.type="string"/>')
    lines.append('  <key id="d8" for="edge" attr.name="txid_label" attr.type="string"/>')
    # yEd rendering keys for visible labels in graph view.
    lines.append('  <key id="yn" for="node" yfiles.type="nodegraphics"/>')
    lines.append('  <key id="ye" for="edge" yfiles.type="edgegraphics"/>')
    lines.append('  <graph id="G" edgedefault="directed">')

    for node in result.get("graph", {}).get("nodes", []):
        node_id = xml_escape(node.get("id"))
        raw_id = str(node.get("id", ""))
        node_type = str(node.get("type", ""))
        if raw_id.startswith("addr:"):
            display_label = raw_id[5:]
        elif raw_id.startswith("tx:"):
            display_label = raw_id[3:]
        else:
            display_label = raw_id
        lines.append(f'    <node id="{node_id}">')
        lines.append(f'      <data key="d0">{xml_escape(node_type)}</data>')
        lines.append(f'      <data key="d1">{xml_escape(node.get("asset"))}</data>')
        lines.append(f'      <data key="d6">{xml_escape(display_label)}</data>')
        lines.append('      <data key="yn">')
        lines.append("        <y:ShapeNode>")
        lines.append(f"          <y:NodeLabel>{xml_escape(display_label)}</y:NodeLabel>")
        lines.append("        </y:ShapeNode>")
        lines.append("      </data>")
        lines.append("    </node>")

    for idx, edge in enumerate(result.get("graph", {}).get("edges", []), start=1):
        src = xml_escape(edge.get("src"))
        dst = xml_escape(edge.get("dst"))
        amount = edge.get("amount")
        asset = str(edge.get("asset", "")).strip()
        txid = str(edge.get("txid", "")).strip()
        amount_label = f"{amount} {asset}".strip()
        edge_label = f"{amount_label}\nTX: {txid}" if txid else amount_label
        lines.append(f'    <edge id="e{idx}" source="{src}" target="{dst}">')
        lines.append(f'      <data key="d2">{xml_escape(edge.get("label"))}</data>')
        lines.append(f'      <data key="d3">{xml_escape(amount)}</data>')
        lines.append(f'      <data key="d4">{xml_escape(txid)}</data>')
        lines.append(f'      <data key="d5">{xml_escape(asset)}</data>')
        lines.append(f'      <data key="d7">{xml_escape(amount_label)}</data>')
        lines.append(f'      <data key="d8">{xml_escape(txid)}</data>')
        lines.append('      <data key="ye">')
        lines.append("        <y:PolyLineEdge>")
        lines.append('          <y:Arrows source="none" target="standard"/>')
        lines.append(f"          <y:EdgeLabel>{xml_escape(edge_label)}</y:EdgeLabel>")
        lines.append("        </y:PolyLineEdge>")
        lines.append("      </data>")
        lines.append("    </edge>")

    lines.append("  </graph>")
    lines.append("</graphml>")

    output_path.write_text("\n".join(lines), encoding="utf-8")


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="aml_crypto_analyzer",
        description="Trace cryptocurrency flows for AML investigations.",
    )
    parser.add_argument("--asset", required=True, choices=sorted(SUPPORTED_ASSETS.keys()), help="Asset symbol")
    parser.add_argument("--txid", help="Starting transaction hash")
    parser.add_argument("--address", help="Starting address")
    parser.add_argument(
        "--max-transactions",
        type=int,
        default=DEFAULT_TRANSACTION_LIMIT,
        help=f"Max transactions to include (default: {DEFAULT_TRANSACTION_LIMIT})",
    )
    parser.add_argument(
        "--all-transactions",
        action="store_true",
        help="Include all transactions available from the provider",
    )
    parser.add_argument("--intel-dir", default="./intel", help="Directory containing intelligence CSV files")
    parser.add_argument("--export-json", default="./exports/trace.json", help="Output JSON path")
    parser.add_argument("--export-xlsx", default="./exports/trace.xlsx", help="Output XLSX path")
    parser.add_argument("--export-graphml", default="./exports/trace.graphml", help="Output GraphML path")
    parser.add_argument("--skip-xlsx", action="store_true", help="Skip XLSX export")
    parser.add_argument("--skip-graphml", action="store_true", help="Skip GraphML export")
    parser.add_argument("--blockcypher-key", default=os.getenv("BLOCKCYPHER_API_KEY"), help="Optional BlockCypher API key")
    parser.add_argument("--blockchair-key", default=os.getenv("BLOCKCHAIR_API_KEY"), help="Optional Blockchair API key")
    parser.add_argument("--coinstats-key", default=os.getenv("COINSTATS_API_KEY"), help="Optional CoinStats API key")
    parser.add_argument("--fiat-currency", default=None, help="Optional ISO currency code (e.g. USD, EUR) for historical conversion")
    parser.add_argument("--print-summary", action="store_true", help="Print findings summary to stdout")

    args = parser.parse_args(argv)
    if not args.txid and not args.address:
        parser.error("Provide either --txid or --address")
    if args.txid and args.address:
        parser.error("Provide only one of --txid or --address")
    if args.all_transactions:
        args.max_transactions = None
    elif args.max_transactions < 1:
        parser.error("--max-transactions must be >= 1, or use --all-transactions")
    return args


def print_summary(result: Dict[str, Any]) -> None:
    meta = result.get("metadata", {})
    print("\n=== Trace Summary ===")
    print(f"Asset: {meta.get('asset')}")
    print(f"Transaction limit: {meta.get('transaction_limit')}")
    print(f"Transactions: {meta.get('transaction_count')}")
    print(f"Nodes: {meta.get('node_count')}")
    print(f"Edges: {meta.get('edge_count')}")
    print(f"Findings: {meta.get('finding_count')}")

    findings = result.get("findings", [])
    if findings:
        print("\nTop findings:")
        for finding in findings[:20]:
            print(
                f"- [{finding.get('level').upper()}] {finding.get('finding_type')} "
                f"| {finding.get('subject')} | {finding.get('detail')}"
            )


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)

    provider = ProviderRouter(
        blockcypher_key=args.blockcypher_key,
        blockchair_key=args.blockchair_key,
        coinstats_key=args.coinstats_key,
    )
    intel = IntelligenceStore(Path(args.intel_dir))
    analyzer = Analyzer(provider=provider, intel=intel)

    started = time.time()
    if args.txid:
        result = analyzer.trace_from_transaction(
            asset=args.asset,
            txid=args.txid,
            max_transactions=args.max_transactions,
        )
    else:
        result = analyzer.trace_from_address(
            asset=args.asset,
            address=args.address,
            max_transactions=args.max_transactions,
        )

    result = enrich_result_with_fiat(result, args.fiat_currency)

    export_json(result, Path(args.export_json))
    if not args.skip_xlsx:
        export_xlsx(result, Path(args.export_xlsx))
    if not args.skip_graphml:
        export_graphml(result, Path(args.export_graphml))

    elapsed = time.time() - started
    if args.print_summary:
        print_summary(result)
    print(f"\nCompleted in {elapsed:.2f}s")
    print(f"JSON: {args.export_json}")
    if not args.skip_xlsx:
        print(f"XLSX: {args.export_xlsx}")
    if not args.skip_graphml:
        print(f"GraphML: {args.export_graphml}")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ProviderError as exc:
        print(f"Provider error: {exc}", file=sys.stderr)
        raise SystemExit(2)
    except KeyboardInterrupt:
        print("Interrupted", file=sys.stderr)
        raise SystemExit(130)
